# -----------------------------------------------------------
# demonstrates how to write ms excel files using python-openpyxl
#
# (C) 2015 Frank Hofmann, Berlin, Germany
# Released under GNU Public License (GPL)
# email frank.hofmann@efho.de
# -----------------------------------------------------------

# include standard modules
from openpyxl import Workbook

# create a workbook
wb = Workbook()

# add a second work sheet at the end
ws1 = wb.create_sheet()

# get active work sheet from the current work sheet
ws = wb.active

# set the work sheet title
ws.title = "The first work sheet"
ws1.title = "The second work sheet"

# get the current work sheet
print "active work sheet: ", ws

# output the names of all the work sheets
print(wb.get_sheet_names())

# fill cells with values
data = {
	"A1": 84,
	"A2": 15,
	"A3": 17,
	"B1": 42,
	"B2": 6,
	"B3": 19
}

for entry in data:
	ws[entry] = data[entry]

# save workbook
wb.save('/tmp/test.xlsx')

