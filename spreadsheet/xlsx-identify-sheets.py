# Demonstrate usage of OpenPyxl to identify sheetnames from an 
# Excel spreadsheet
#
# (C) 2024 Frank Hofmann, Freiburg, Germany <info@efho.de>
# Published under GNU Public License (GPL)
#
# Inspired by examples from:
# https://realpython.com/openpyxl-excel-spreadsheets-python/

from openpyxl import Workbook, load_workbook
filename = "hello_world.xlsx"

workbook = load_workbook(filename)
print(workbook.sheetnames)

